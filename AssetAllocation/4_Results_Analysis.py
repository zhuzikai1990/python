# Results Analysis.py

import numpy as np
import pandas as pd

# deltatime calculation
import datetime
from datetime import date
import time

# Plot Package
import matplotlib.pyplot as plt
import seaborn as sns
sns.set()

# Excel Handling
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.styles import NamedStyle, Font, Border, Side, PatternFill
from openpyxl.drawing.image import Image

# files checking
import os
import re
import glob

# Get into those file with results
result_file = 'Result_1113'
os.chdir(os.path.dirname(__file__) + '//' + result_file)
root = os.getcwd()


def listfiles(path):
    path = path.replace("\\", "/")
    mlist = os.listdir(path)
    pts = np.array([])
    for m in mlist:
        mpath = os.path.join(path, m)
        if os.path.isfile(mpath):
            pt = os.path.abspath(mpath)
            # print pt.decode("gbk").encode("utf-8") #会报错
        else:
            pt = os.path.abspath(mpath)
        pts = np.append(pts, pt)
    return pts


def change_to_string(array):
    return np.char.decode(array.astype(np.string_))


def main():
    global root
    files = glob.glob('*')
    pathes = [root + '//' + file for file in files]

    file_arrays = np.array([])
    for path in pathes:
        file_array = listfiles(path)
        file_arrays = np.append(file_arrays, file_array)

    # get cumulative return picture and backtesting result id
    cum_files_id = [True if 'BT_Cum_Return' in file else False for file in file_arrays]
    backtesting_files_id = [
        True if 'BT_Results' in file else False for file in file_arrays
    ]

    #1_2002-1_2018-7_0_0_delta0.1_EWMA0.9943_BF_AssetAllocation_Backtesting_Results

    # get cumulative return picture and backtesting result content
    cum_files = file_arrays[cum_files_id]
    backtesting_files = file_arrays[backtesting_files_id]

    # get junson portfolio result
    curr_result = pd.read_csv(
        backtesting_files[0], index_col=0).drop(
            ['MSCI Developed', 'Bond Aggregate', 'Benchmark'], axis=0)
    curr_results = pd.DataFrame([])

    # get backtesting result and corresponding parameter settings and concate them
    for backtesting_file in backtesting_files:
        print(backtesting_file)
        curr_result = pd.read_csv(
            backtesting_file, index_col=0).drop(
                ['MSCI Developed', 'Bond Aggregate', 'Benchmark'], axis=0)

        attributes = backtesting_file.split('\\')[-2].split('_')
        try:
            if int(attributes[6]) == 0:
                delta_con = 0
                delta_val = 0
        except:
            delta_con = attributes[8][:-3]
            delta_val = attributes[8][-3:]

        try:
            if int(attributes[7]) == 0:
                EWMA_mtd = 0
                EWMA_val = 0
        except:
            EWMA_mtd = attributes[9][:4]
            EWMA_val = attributes[9][4:]

        values = {
            'metric': attributes[0],
            'opt_len': attributes[1],
            'ST': attributes[2],
            'ET': attributes[3],
            'asset_con': attributes[4],
            'group_con': attributes[5],
            'growth_ratio':attributes[6],
            'defence_ratio':attributes[7],
            'delta_con': delta_con,
            'delta_val': delta_val,
            'EWMA_mtd': EWMA_mtd,
            'EWMA_val': EWMA_val,
            'Fill_mtd': attributes[10],
        }

        att_df = pd.DataFrame(data=values, index=['Junson Portfolio'])

        curr_result = pd.concat([curr_result, att_df], axis=1)
        curr_results = pd.concat([curr_results, curr_result], axis=0)
    attribute_cols = list(values.keys())
    curr_results.reset_index(inplace=True)

    backtesting_files_str = change_to_string(backtesting_files)
    backtesting_files_split = np.char.split(backtesting_files_str, '\\')

    parameters = np.array([])
    for i in backtesting_files_split:
        parameter = i[-2]
        parameters = np.append(parameters, parameter)

    parameters = pd.Series(parameters)
    parameters_df = pd.DataFrame(parameters, columns=['parameter'])
    parameters_df.reset_index(inplace=True)
    parameters_df.rename(columns={'index': 'level_0'}, inplace=True)
    curr_results.reset_index(inplace=True)
    performance = pd.merge(curr_results, parameters_df)

    # find best indicator id
    performance.drop(['level_0'], axis=1, inplace=True)
    performance.drop(
        ['index', 'MaxDrowndown_Start', 'MaxDrownDown_End'],
        axis=1,
        inplace=True)
    performance[
        'Standard Deviation(%)'] = -performance['Standard Deviation(%)']
    performance['MaxDrowndown(%)'] = -performance['MaxDrowndown(%)']
    performance['Monthly Turnover(%)'] = -performance['Monthly Turnover(%)']
    attributes_df = performance[attribute_cols]
    performance.drop(attribute_cols, axis=1, inplace=True)
    best_performances_s = performance.apply(max)
    best_performances_v = best_performances_s.values
    best_performances_i = best_performances_s.index

    pictures = []
    print(
        'Best Performance Parameter Setting in different dimension: Annual Return, Sharpe Ratio, MaxDrownDown, Excess Return.'
    )

    for i in range(len(best_performances_v) - 1):
        if i == 1:
            continue
        else:
            print(best_performances_i[i])
            id_ = performance[performance.iloc[:, i] == best_performances_v[
                i]].index.values[0]
            picture = cum_files[performance.index[id_]]
            pictures.append(picture)
            #img=Image.open(picture)
            #plt.figure(parameters[performance.index[id_]],figsize = (15,9))
            #plt.imshow(img)
            #plt.axis('off')
            #plt.show()
            #print('------------------------------------------------------')

    pictures = set(pictures)

    performance[
        'Standard Deviation(%)'] = -performance['Standard Deviation(%)']
    performance['MaxDrowndown(%)'] = -performance['MaxDrowndown(%)']
    performance['Monthly Turnover(%)'] = -performance['Monthly Turnover(%)']
    performance = pd.concat([performance, attributes_df], axis=1)
    performance.to_excel('P_Performance.xlsx')

    # add picture into excel
    wb = load_workbook("P_Performance.xlsx")
    ws = wb.active

    highlight = NamedStyle(name="highlight")
    highlight.font = Font(bold=True, size=20, color="ff0100")
    highlight.fill = PatternFill("solid", fgColor="DDDDDD")
    bd = Side(style='thick', color="000000")  #边框颜色及粗细
    highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)  #边框 上下左右

    pictures = list(pictures)
    for i in range(len(pictures)):
        picture = pictures[i]
        print(picture)
        name = picture.split('\\')[-2]
        ws["X" + str(1 + i * 50)] = name
        ws["X" + str(1 + i * 50)].style = highlight
        img = Image(picture)
        ws.add_image(img, "X" + str(3 + i * 50))
    wb.save('P_Performance_processed.xlsx')


if __name__ == '__main__':
    main()
